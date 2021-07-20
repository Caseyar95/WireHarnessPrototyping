#! Python 3
# - Copy and Paste Ranges using OpenPyXl library
 
import openpyxl
 
#Prepare the spreadsheets to copy from and paste too.
 
#File to be copied
wb = openpyxl.load_workbook("foo.xlsx") #Add file name
sheet = wb["foo"] #Add Sheet name
 
#File to be pasted into
#template = openpyxl.load_workbook("foo2.xlsx") #Add file name
#temp_sheet = wb["foo2"] #Add Sheet name
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
        
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            cell = sheetReceiving.cell(row = i, column = j)
            cell.value = copiedData[countRow][countCol]
            if countCol == 3:
                ########
            countCol += 1
        countRow += 1
        
def newRow(row, curSheet):
    curSheet.insert_rows(row)
    curSheet.save("foo.xlsx")

def createData():
    print("Processing...")
    selectedRange = copyRange(1,1,18,1,sheet) #Change the 4 number values
    rowCount = sheet.max_row
    while i < rowCount:
        x = i+1
        cell = sheet.cell(row = i, column = 1).value
        cell2 = sheet.cell(row = x, column = 1).value
        if cell != cell2:
            newRow(x, sheet)
            ############
    pastingRange = pasteRange(1,3,4,15,sheet,selectedRange) #Change the 4 number values
    #You can save the template as another file to create a new file here too.s
    wb.save("foo.xlsx")
    print("Range copied and pasted!")