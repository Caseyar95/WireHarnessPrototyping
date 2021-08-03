import openpyxl as xl

print ("Loading...")
#opening workbook, loading sheetnames
filename ="ARTOS_Wire_Cut_List_CNH_47714208R_7_19_21.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# calculate total number of rows
mr = ws1.max_row

print ("Calculating...")

for i in range (2, mr + 1):    #loop through rows, if value is target connector, update target column
    c = ws1.cell(row = i, column = 1).value
    z = ws1.cell(row = i, column = 19).value
    y = ws1.cell(row = i, column = 14).value
    int(y)
    #print (type(y))
    if c in {"X-149"}:
        ws1.cell(row = i, column = 14).value = ws1.cell(row = i, column = 14).value + (-8)
        #print (c) ###TEST
    if z in {"X-149"}:
        ws1.cell(row = i, column = 14).value = ws1.cell(row = i, column = 14).value + (-8)

print ("Finished")
#saving file
wb1.save(str(filename))
