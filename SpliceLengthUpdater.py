import openpyxl as xl

#opening workbook, loading sheetnames
filename ="ARTOS_Wire_Cut_List_CNH_47714208R_7_19_21.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[1]
ws2 = wb1.worksheets[2]

print ("Loading Splice Wire-Length Updater")

# calculate total number of rows
mr = ws1.max_row
mr2 = ws2.max_row

print("Updating wire lengths...")

for y in range (1, mr2 + 1):
    v = ws2.cell(row = y, column = 3).value
    #print (v) ###TEST
    for i in range (1, mr + 1):
        c = ws1.cell(row = i, column = 6).value
        #print (c) ###TEST
        if (c == v):
            #print (c, " ", v)  ###TEST
            ws2.cell(row = y, column = 11).value = ws1.cell(row = i, column = 15).value

print ("Saving...")

#saving file
wb1.save(str(filename))

print ("Finished")
