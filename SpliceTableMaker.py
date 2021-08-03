import openpyxl as xl
from copy import copy


#opening workbook, loading sheetnames
filename ="4208Test.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[1]

# calculate total number of rows
mr = ws1.max_row
i = 2
while i < mr + 1:    #loop through rows, if value is target connector, update target column
    h = i + 1
    c = ws1.cell(row = i, column = 1).value
    z = ws1.cell(row = h, column = 1).value
    if z != c:
        ws1.insert_rows(h)
        for y in range(1, 18):
            ws1.cell(row = h, column = y).value = ws1.cell(row = 1, column = y).value
            ws1.cell(row = h, column = y)._style = copy(ws1.cell(row = 1, column = y))
        i= i+2
    i = i+1
        

#saving file
wb1.save(str(filename))
