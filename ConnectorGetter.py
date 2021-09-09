import openpyxl as xl

#opening workbook, loading sheetnames
filename ="C:/Users/Harness Laptop/Documents/Casey/CNH 4204/ARTOS_Master_Wire_Cut_List _47714204_8_19_2021.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# calculate total number of rows
mr = ws1.max_row
conn = {"X-012"}
count = 0
connectors = []
ConnList = []

for i in range (1, mr + 1):    #loop through rows, if value is target connector, get conn, pin, ID, AWG
    c = ws1.cell(row = i, column = 1).value
    z = ws1.cell(row = i, column = 20).value
    if c in conn:
        
        print(ws1.cell(row = i, column = 1).value," ",ws1.cell(row = i, column = 2).value, " ", ws1.cell(row = i, column = 6).value,
              " ",ws1.cell(row = i, column = 10).value," ") #ws1.cell(row = i, column = 9).value, "\n")
        count = count + 1
        #print (c) ###TEST
    if z in conn:
        print(ws1.cell(row = i, column = 20).value," ",ws1.cell(row = i, column = 21).value," ", ws1.cell(row = i, column = 6).value,
              " ",ws1.cell(row = i, column = 10).value," ") #ws1.cell(row = i, column = 9).value, "\n")
        count = count + 1
    
print (count)

#saving file
#wb1.save(str(filename))
