#! Python 3
# - Copy and Paste Ranges using OpenPyXl library
 
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
 
#Prepare the spreadsheets to copy from and paste too.
 
#File to be copied
wb = openpyxl.load_workbook("ARTOS_Wire_Cut_List - JASPER 12072046_7_15_21.xlsx") #Add file name
sheet = wb["test"] #Add Sheet name
 
#File to be pasted into
#template = openpyxl.load_workbook("foo2.xlsx") #Add file name
#temp_sheet = wb["foo2"] #Add Sheet name
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
        
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            cell = sheetReceiving.cell(row = i, column = j)
            cell.value = copiedData[countRow][countCol]
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text= True, vertical="center")
            cell.border = black_border
            #if countCol == 3:
                ########
            countCol += 1
        countRow += 1
        
def newRow(row, curSheet):
    curSheet.insert_rows(row)
    curSheet.insert_rows(row)
    #curSheet.save("foo.xlsx")

#def createData():
print("Processing...")
selectedRange = copyRange(1,1,18,2,sheet) #Change the 4 number values
rowCount = sheet.max_row
i = 1
thin = Side(border_style="thin", color= "303030")
black_border = Border(top=thin, left=thin, right=thin, bottom=thin)

while i < rowCount:
    x = i+1
    y = x+1
    cell = sheet.cell(row = i, column = 1).value
    cell2 = sheet.cell(row = x, column = 1).value
    if cell != cell2:
        newRow(x, sheet)
        pastingRange = pasteRange(1,x,18,y,sheet,selectedRange) #Change the 4 number values
        
        i = y+2
    else:
        i = x
            ############
    
    #You can save the template as another file to create a new file here too.s
wb.save("ARTOS_Wire_Cut_List - JASPER 12072046_7_15_21.xlsx")
print("Range copied and pasted!")
