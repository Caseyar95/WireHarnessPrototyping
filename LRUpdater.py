import openpyxl as xl

#opening workbook, loading sheetnames
filename ="ARTOS_Master_Wire_Cut_List_CNH 48018092_E1.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]
ws2 = wb1.worksheets[1]

print ("Loading Orientation Updater")

# calculate total number of rows
mr = ws1.max_row
mr2 = ws2.max_row

print("Updating splice orientations...")

for y in range (1, mr2 + 1):
    v = ws2.cell(row = y, column = 7).value   #load ID
    z = ws2.cell(row = y, column = 1).value   #load Splice ID
    #print (v) ###TEST
    for i in range (1, mr + 1):
        c = ws1.cell(row = i, column = 6).value   #load ID from master
        x = ws1.cell(row = i, column = 1).value   #load splice/conn from master
        n = ws1.cell(row = i, column = 19).value   #load other s/c from master
        #print (c) ###TEST
        if (c == v):   #compare ID's
            if (z == x):   #compare s/c ID's
                #print (c, " ", v)  ###TEST
                ws1.cell(row = i, column = 2).value = ws2.cell(row = y, column = 2).value
            if (z == n):   #oompare other s/c ID's
                ws1.cell(row = i, column = 20).value = ws2.cell(row = y, column = 2).value

print ("Saving...")

#saving file
wb1.save(str(filename))

print ("Finished")
